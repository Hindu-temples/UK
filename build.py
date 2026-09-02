#!/usr/bin/env python3
"""
Hindu Temples UK — site builder.

    python build.py

One command does everything:
  1. If temples.xlsx is present (and openpyxl installed), rebuilds temples.json from it.
  2. Generates every page: index, directory, 20 city pages, temple-map, about, 404.
  3. Generates sitemap.xml, robots.txt, site.webmanifest, favicon.svg,
     og-image.png and apple-touch-icon.png.

Edit temples.xlsx  ->  python build.py  ->  commit. That's the whole workflow.
"""
import json, html, math, os, sys, datetime

BASE   = "https://hindu-temples.uk"
EMAIL  = "hindu.temples.uk@gmail.com"
TODAY  = datetime.date.today().isoformat()
MAILTO = ("mailto:" + EMAIL +
  "?subject=Temple%20submission%20%E2%80%93%20hindu-temples.uk"
  "&amp;body=Temple%20name%3A%0AAddress%20%2F%20postcode%3A%0A"
  "What%27s%20new%2C%20closed%2C%20or%20wrong%3A%0AOpening%20hours%20(if%20known)%3A%0A"
  "Website%20%2F%20social%20links%3A%0A")

TRAD_COLOR = {"BAPS":"#B8922F","ISKCON":"#2E4A6B","Saiva":"#BE3524",
              "Swaminarayan":"#C56A2A","General":"#0E6B67"}
TRAD_LABEL = {"BAPS":"BAPS Swaminarayan","ISKCON":"ISKCON / Vaishnava","Saiva":"Saiva / Tamil",
              "Swaminarayan":"Swaminarayan (other)","General":"Sanatan / community"}
REGIONS = ["London","South East","South West","East of England","East Midlands","West Midlands",
           "Yorkshire and the Humber","North West","North East","Scotland","Wales","Northern Ireland"]

# name, slug, lat, lng, radius(mi), zoom, blurb
CITIES = [
 ("London","london",51.509,-0.126,15,10,"London is home to the largest Hindu population in the UK, with major mandirs across the capital from Neasden to Southall. The borough of Harrow has the highest proportion of Hindu residents of any local authority in England and Wales."),
 ("Leicester","leicester",52.6369,-1.1398,12,12,"Leicester has one of the most established Hindu communities in Britain, centred on the famous Golden Mile along Belgrave Road, and hosts one of the largest Diwali celebrations outside India."),
 ("Birmingham","birmingham",52.4862,-1.8904,12,11,"Birmingham and the wider West Midlands host a large, diverse Hindu community. Nearby Tividale is home to the Shri Venkateswara (Balaji) Temple, one of the largest Hindu temple complexes in Europe."),
 ("Coventry","coventry",52.4068,-1.5197,8,12,"Coventry has a long-established Hindu community, with a cluster of temples spanning the Sanatan, Swaminarayan, ISKCON and South Indian Saiva traditions."),
 ("Wolverhampton","wolverhampton",52.5870,-2.1288,8,12,"Wolverhampton and the Black Country are home to a settled Hindu community, with mandirs across Wolverhampton, Walsall and Bilston."),
 ("Manchester","manchester",53.4808,-2.2426,9,11,"Greater Manchester's Hindu community is served by long-standing mandirs across the city and surrounding towns, from Withington to Whalley Range."),
 ("Bolton","bolton",53.5769,-2.4282,7,12,"Bolton has a notable Gujarati Hindu community, with several Swaminarayan and Sanatan temples."),
 ("Preston","preston",53.7632,-2.7031,8,12,"Preston is home to the Gujarat Hindu Society, one of the largest Hindu temples in Lancashire."),
 ("Liverpool","liverpool",53.4084,-2.9916,10,11,"Liverpool and Merseyside are served by a mix of North Indian and South Indian Tamil Hindu temples."),
 ("Leeds","leeds",53.8008,-1.5491,9,12,"Leeds has a growing Hindu community, with a central mandir and a BAPS Swaminarayan temple in Burley."),
 ("Bradford","bradford",53.7960,-1.7594,7,12,"Bradford's Hindu community is served by the long-established Hindu Cultural Society."),
 ("Nottingham","nottingham",52.9548,-1.1581,8,12,"Nottingham's Hindu temple and community centre on Carlton Road serves the wider East Midlands city."),
 ("Derby","derby",52.9225,-1.4746,8,12,"Derby's main mandir, the Geeta Bhawan, has served the city's Hindu community for decades."),
 ("Luton","luton",51.8787,-0.4200,7,12,"Luton has a well-established Hindu community, with a large community temple and a BAPS Swaminarayan mandir."),
 ("Bedford","bedford",52.1360,-0.4667,7,12,"Bedford's Hindu community is served by a community temple and cultural trust."),
 ("Watford","watford",51.656,-0.396,8,12,"Watford and the surrounding Hertfordshire villages are home to Bhaktivedanta Manor, the ISKCON estate famously donated by George Harrison in 1973."),
 ("Peterborough","peterborough",52.5695,-0.2405,7,12,"Peterborough is served by the Bharat Hindu Samaj, long the city's main Hindu temple."),
 ("Cardiff","cardiff",51.4816,-3.1791,10,12,"Cardiff has the largest Hindu community in Wales, with a Swaminarayan temple and a Sanatan Dharma community centre."),
 ("Glasgow","glasgow",55.8642,-4.2518,10,11,"Glasgow is home to Scotland's largest Hindu community, with several mandirs including one of the country's oldest."),
 ("Slough","slough",51.509,-0.595,6,12,"Slough is home to one of the first purpose-built Hindu temples in Britain, serving a large community west of London."),
]

FEATURED = [
 ("BAPS Shri Swaminarayan Mandir","Neasden","london","Europe's first traditional stone mandir (1995), hand-carved in marble and limestone."),
 ("Shri Venkateswara (Balaji) Temple","Tividale","birmingham","One of the largest Hindu temple complexes in Europe, in the South Indian style."),
 ("Bhaktivedanta Manor","Aldenham","watford","The ISKCON (Hare Krishna) estate donated by George Harrison in 1973."),
 ("Bharat Hindu Samaj Mandir","Peterborough","peterborough","Long the main mandir serving Peterborough and the surrounding area."),
 ("Shree Sanatan Mandir","Leicester","leicester","A carved-limestone landmark serving Leicester's Golden Mile community."),
 ("Shree Ghanapathy Temple","Wimbledon","london","A leading South London Saiva temple, dedicated to Lord Ganesha."),
]

FAQS = [
 ("How many Hindu temples are there in the UK?",
  'This directory lists <b>__N__ verified Hindu temples and mandirs</b> across England, Scotland, Wales and Northern Ireland, each checked against a real street address. National bodies such as the National Council of Hindu Temples UK estimate several hundred places of Hindu worship in total once smaller community shrines and home gatherings are included.'),
 ("Which UK city has the most Hindu temples?",
  'London has by far the most, with <a href="london.html">__NLON__ temples listed</a> across the capital. <a href="leicester.html">Leicester</a> follows with __NLEI__, clustered around the Golden Mile on Belgrave Road \u2014 home to one of the largest Diwali celebrations outside India.'),
 ("What is the largest Hindu temple in the UK?",
  'The <b>BAPS Shri Swaminarayan Mandir in Neasden</b>, London \u2014 opened in 1995 \u2014 was Europe\u2019s first traditional stone mandir and remains the best known. The <b>Shri Venkateswara (Balaji) Temple</b> in Tividale, near Birmingham, is among the largest Hindu temple complexes in Europe.'),
 ("How do I find my nearest Hindu temple?",
  'Type your postcode into the search box at the top of this page. The map ranks every temple by distance, draws sightlines to the four nearest, and gives you one-tap directions, phone numbers and websites where available.'),
 ("Are temple opening times listed?",
  'Opening hours are shown where they have been verified, and more are added as temples confirm them. Times can change for festivals and private events, so it\u2019s always worth checking with the temple before travelling.'),
 ("A temple is missing or has closed \u2014 how do I report it?",
  'Email <a href="__MAILTO__">' + EMAIL + '</a> with the temple\u2019s name and address and what needs changing. Every submission is checked before the map is updated \u2014 see <a href="about.html">how this directory is maintained</a>.'),
]

# ---------------------------------------------------------------- data
def s(v): return "" if v is None else str(v).strip()

def load_from_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb["Temples"] if "Temples" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    header = [s(h) for h in rows[0]]
    def col(name):
        return header.index(name) if name in header else None
    idx = {k: col(k) for k in ["Temple Name","Area","Region","County","Latitude","Longitude",
           "Address","Phone","Tradition","Notes","Featured","Status","Opening hours",
           "Website","Facebook","Instagram","Last updated"]}
    for req in ("Temple Name","Latitude","Longitude"):
        if idx[req] is None: sys.exit("ERROR: column '%s' missing from %s" % (req, path))
    def get(row, k):
        i = idx[k]; return row[i] if i is not None and i < len(row) else None
    labels = {v.lower(): k for k, v in TRAD_LABEL.items()}
    out, skipped = [], []
    for n, row in enumerate(rows[1:], start=2):
        name = s(get(row, "Temple Name"))
        if not name: continue
        try:
            lat = round(float(get(row, "Latitude")), 4); lng = round(float(get(row, "Longitude")), 4)
        except (TypeError, ValueError):
            skipped.append((n, name)); continue
        traw = s(get(row, "Tradition"))
        trad = traw if traw in TRAD_LABEL else labels.get(traw.lower(), "General")
        rec = {"name": name, "area": s(get(row, "Area")), "region": s(get(row, "Region")),
               "county": s(get(row, "County")), "lat": lat, "lng": lng,
               "addr": s(get(row, "Address")), "phone": s(get(row, "Phone")),
               "trad": trad, "note": s(get(row, "Notes")),
               "hl": s(get(row, "Featured")).lower() in ("yes","y","true","1")}
        for xk, jk in (("Status","status"),("Opening hours","hours"),("Website","url"),
                       ("Facebook","fb"),("Instagram","ig"),("Last updated","updated")):
            v = s(get(row, xk))
            if v and not (jk == "status" and v.lower() == "open"): rec[jk] = v
        out.append(rec)
    if skipped:
        print("Skipped %d row(s) with no coordinates:" % len(skipped))
        for n, nm in skipped: print("  row %d: %s" % (n, nm))
    return out

def load_festivals():
    """Read the Festivals sheet. Past dates are dropped at build time; the browser
    drops any that expire between rebuilds (see EXPIRY_JS)."""
    if not os.path.exists("temples.xlsx"): return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    wb = load_workbook("temples.xlsx", data_only=True)
    if "Festivals" not in wb.sheetnames: return []
    ws = wb["Festivals"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    hdr = [s(h) for h in rows[0]]
    ix = {k: (hdr.index(k) if k in hdr else None)
          for k in ["Date","End date","Festival","Temple Name","Area","Time","Details","Source","Checked"]}
    def cell(row, k):
        i = ix[k]
        v = row[i] if i is not None and i < len(row) else None
        if hasattr(v, "strftime"): return v.strftime("%Y-%m-%d")
        return s(v)
    out = []
    for row in rows[1:]:
        d = cell(row, "Date"); f = cell(row, "Festival")
        if not d or not f: continue
        end = cell(row, "End date") or d
        if end < TODAY: continue                      # build-time expiry
        out.append({"date": d, "end": end, "fest": f,
                    "temple": cell(row, "Temple Name"), "area": cell(row, "Area"),
                    "time": cell(row, "Time"), "details": cell(row, "Details"),
                    "source": cell(row, "Source")})
    out.sort(key=lambda x: x["date"])
    return out

def pretty_date(iso, end=None):
    import datetime as _dt
    try: a = _dt.date.fromisoformat(iso)
    except ValueError: return iso
    fmt = lambda x: "%d %s %d" % (x.day, x.strftime("%B"), x.year)
    if end and end != iso:
        try:
            b = _dt.date.fromisoformat(end)
            if b.month == a.month:
                return "%d\u2013%d %s %d" % (a.day, b.day, a.strftime("%B"), a.year)
            return "%s \u2013 %s" % (fmt(a), fmt(b))
        except ValueError: pass
    return fmt(a)

def load_data():
    if os.path.exists("temples.xlsx"):
        try:
            data = load_from_xlsx("temples.xlsx")
            json.dump(data, open("temples.json","w",encoding="utf-8"),
                      ensure_ascii=False, separators=(",",":"))
            print("data: rebuilt temples.json from temples.xlsx (%d temples)" % len(data))
            return data
        except ImportError:
            print("data: openpyxl not installed - using existing temples.json")
    data = json.load(open("temples.json", encoding="utf-8"))
    print("data: loaded temples.json (%d temples)" % len(data))
    return data

T = load_data()
for i, t in enumerate(T): t["id"] = i

def miles(a,b,c,d):
    R=3958.8; r=math.pi/180
    x=math.sin((c-a)*r/2)**2+math.cos(a*r)*math.cos(c*r)*math.sin((d-b)*r/2)**2
    return R*2*math.asin(math.sqrt(x))

for t in T:
    if t["region"] == "London":
        t["_city"] = "london"; continue
    best, bd = None, 1e9
    for nm, slug, la, lo, rad, z, bl in CITIES:
        if slug == "london": continue
        d = miles(t["lat"], t["lng"], la, lo)
        if d <= rad and d < bd: bd, best = d, slug
    t["_city"] = best

BY_CITY = {c[1]: [] for c in CITIES}
for t in T:
    if t["_city"]: BY_CITY[t["_city"]].append(t)
for slug in BY_CITY:
    BY_CITY[slug].sort(key=lambda x: (0 if x.get("hl") else 1, x["name"]))

FEST = load_festivals()
FEST_NATIONAL = [f for f in FEST if not f["temple"]]
FEST_BY_TEMPLE = {}
for _f in FEST:
    if _f["temple"]:
        FEST_BY_TEMPLE.setdefault((_f["temple"].lower(), _f["area"].lower()), []).append(_f)
print("festivals: %d upcoming (%d national, %d temple-specific)" % (
    len(FEST), len(FEST_NATIONAL), len(FEST) - len(FEST_NATIONAL)))

N = len(T)
N_LON = len(BY_CITY["london"]); N_LEI = len(BY_CITY["leicester"])
REGION_COUNT = {r: sum(1 for t in T if t["region"] == r) for r in REGIONS}

# ---------------------------------------------------------------- helpers
def e(x): return html.escape(str(x or ""), quote=True)
def sub(tpl, **kw):
    for k, v in kw.items(): tpl = tpl.replace("__%s__" % k, v)
    return tpl
def gdir(t): return "https://www.google.com/maps/dir/?api=1&amp;destination=%s,%s" % (t["lat"], t["lng"])
def ld(obj): return '<script type="application/ld+json">%s</script>' % json.dumps(
    obj, ensure_ascii=False, separators=(",", ":"))

FAVICON_SVG = ('<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 64 64">'
 '<rect width="64" height="64" rx="14" fill="#BE3524"/>'
 '<path d="M32 8 L44 38 H20 Z" fill="#FBF3E4"/>'
 '<rect x="16" y="38" width="32" height="12" fill="#FBF3E4"/>'
 '<rect x="28" y="41" width="8" height="9" fill="#BE3524"/>'
 '<circle cx="32" cy="6.5" r="2.5" fill="#E8A13A"/></svg>')

BRAND_SVG = ('<svg viewBox="0 0 64 64" aria-hidden="true">'
 '<rect width="64" height="64" rx="14" fill="#BE3524"/>'
 '<path d="M32 8 L44 38 H20 Z" fill="#FBF3E4"/>'
 '<rect x="16" y="38" width="32" height="12" fill="#FBF3E4"/>'
 '<rect x="28" y="41" width="8" height="9" fill="#BE3524"/>'
 '<circle cx="32" cy="6.5" r="2.5" fill="#E8A13A"/></svg>')

HAS_OG = False  # set true if og-image.png generated

def head(title, desc, canon, active="", leaflet=False, noindex=False, extra=""):
    og = ('<meta property="og:image" content="%s/og-image.png"/>'
          '<meta name="twitter:card" content="summary_large_image"/>' % BASE) if HAS_OG else \
         '<meta name="twitter:card" content="summary"/>'
    return sub("""<!DOCTYPE html>
<html lang="en-GB">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>__TITLE__</title>
<meta name="description" content="__DESC__"/>
<link rel="canonical" href="__CANON__"/>__ROBOTS__
<meta name="theme-color" content="#FAF6EE"/>
<meta property="og:type" content="website"/>
<meta property="og:site_name" content="Hindu Temples UK"/>
<meta property="og:title" content="__TITLE__"/>
<meta property="og:description" content="__DESC__"/>
<meta property="og:url" content="__CANON__"/>
__OG__
<link rel="icon" type="image/svg+xml" href="favicon.svg"/>
<link rel="apple-touch-icon" href="apple-touch-icon.png"/>
<link rel="manifest" href="site.webmanifest"/>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
__TILES__<link href="https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,400;9..144,500;9..144,600&family=Inter:wght@400;500;600&family=IBM+Plex+Mono:wght@400;500&display=swap" rel="stylesheet">
__LEAFLET__<link rel="stylesheet" href="site.css"/>
__EXTRA__
<!-- analytics: paste your snippet here -->
</head>
<body>
<a class="skip-link" href="#main">Skip to content</a>
""", TITLE=e(title), DESC=e(desc), CANON=canon,
     ROBOTS='\n<meta name="robots" content="noindex,nofollow"/>' if noindex else "",
     OG=og,
     TILES='<link rel="preconnect" href="https://tile.openstreetmap.org"/>\n' if leaflet else "",
     LEAFLET='<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/leaflet/1.9.4/leaflet.min.css"/>\n' if leaflet else "",
     EXTRA=extra)

def header_html(active=""):
    def a(href, label, key):
        cur = ' aria-current="page"' if key == active else ""
        return '<a href="%s"%s>%s</a>' % (href, cur, label)
    return ('<header class="site-header"><div class="bar">'
            '<a class="brand" href="index.html">%s<b>Hindu Temples UK</b></a>'
            '<nav class="site-nav" aria-label="Main">%s%s%s%s%s%s</nav>'
            '</div><div class="kalash"></div></header>\n<main id="main">' % (
        BRAND_SVG,
        a("index.html", "Find a temple", "home"),
        a("directory.html", "Directory", "directory"),
        a("temple-map.html", "Temple map", "map"),
        a("visiting-a-hindu-temple.html", "Visiting", "visiting"),
        a("hinduism-in-britain.html", "History", "history"),
        a("about.html", "About", "about")))

def footer_html():
    cities = " ".join('<a href="%s.html">%s</a> ·' % (c[1], e(c[0])) for c in CITIES).rstrip(" ·")
    return sub("""</main>
<footer class="site-footer">
  <div class="cols">
    <div class="f-brand">
      <b>Hindu Temples UK</b>
      <p>A free, community-maintained directory of Hindu temples and mandirs across the United Kingdom — every listing verified against a real street address.</p>
    </div>
    <div>
      <p class="f-h">Explore</p>
      <ul class="f-links">
        <li><a href="index.html">Find your nearest temple</a></li>
        <li><a href="directory.html">Full A–Z directory</a></li>
        <li><a href="temple-map.html">Map of every UK temple</a></li>
        <li><a href="visiting-a-hindu-temple.html">Visiting a temple</a></li>
        <li><a href="temple-traditions.html">Temple traditions</a></li>
        <li><a href="hinduism-in-britain.html">Hinduism in Britain</a></li>
        <li><a href="about.html">About &amp; data sources</a></li>
        <li><a href="__MAILTO__">Report or add a temple</a></li>
      </ul>
    </div>
    <div>
      <p class="f-h">Temples by city</p>
      <p class="f-cities">__CITIES__</p>
    </div>
  </div>
  <div class="f-meta"><div class="in">
    <span>© __YEAR__ hindu-temples.uk · __N__ temples · Updated __TODAY__</span>
    <span><a href="__MAILTO__">__EMAIL__</a></span>
  </div></div>
</footer>
""" + EXPIRY_JS + """
</body>
</html>""", MAILTO=MAILTO, CITIES=cities, YEAR=TODAY[:4], N=str(N), TODAY=TODAY, EMAIL=EMAIL)

def temple_article(t):
    loc = " · ".join(x for x in [t["area"], t["county"]] if x)
    closed = ' <span class="tag closed">Closed</span>' if str(t.get("status","")).lower()=="closed" else ""
    meta = ""
    if t.get("hours"): meta += '<p class="t-meta">Opening hours: %s</p>' % e(t["hours"])
    if t.get("note"):  meta += '<p class="t-meta">%s</p>' % e(t["note"])
    links = ['<a href="%s" target="_blank" rel="noopener">Directions →</a>' % gdir(t)]
    if t.get("url"):   links.append('<a href="%s" target="_blank" rel="noopener">Website</a>' % e(t["url"]))
    if t.get("phone"): links.append('<a href="tel:%s">%s</a>' % (e(t["phone"].replace(" ","")), e(t["phone"])))
    return ('<article class="temple" id="t-%d">'
            '<h3>%s%s</h3>'
            '<p class="t-sub"><span class="tag">%s</span> %s%s</p>'
            '<p class="t-addr">%s</p>%s%s'
            '<p class="t-links">%s</p></article>' % (
        t["id"], e(t["name"]), " ★" if t.get("hl") else "",
        e(TRAD_LABEL.get(t["trad"], t["trad"])), e(loc), closed, e(t["addr"]), meta,
        temple_fest_html(t), " ".join(links)))

def ld_place(t, compact=False):
    p = {"@type": "HinduTemple", "name": t["name"],
         "address": {"@type": "PostalAddress", "streetAddress": t["addr"],
                     "addressLocality": t["area"], "addressRegion": t["county"], "addressCountry": "GB"}}
    if not compact:
        p["geo"] = {"@type": "GeoCoordinates", "latitude": t["lat"], "longitude": t["lng"]}
        if t.get("phone"): p["telephone"] = t["phone"]
        if t.get("url"): p["url"] = t["url"]
    return p

def ld_itemlist(temples, name, compact=False):
    return {"@context": "https://schema.org", "@type": "ItemList", "name": name,
            "numberOfItems": len(temples),
            "itemListElement": [{"@type": "ListItem", "position": i, "item": ld_place(t, compact)}
                                for i, t in enumerate(temples, 1)]}

def ld_breadcrumb(trail):
    return {"@context": "https://schema.org", "@type": "BreadcrumbList",
            "itemListElement": [{"@type": "ListItem", "position": i+1, "name": n, "item": BASE + "/" + u}
                                for i, (n, u) in enumerate(trail)]}

EXPIRY_JS = """<script>
/* Dates remove themselves once passed, so nothing stale shows between rebuilds. */
(function(){
  var t=new Date(),today=t.getFullYear()+'-'+String(t.getMonth()+1).padStart(2,'0')+'-'+String(t.getDate()).padStart(2,'0');
  document.querySelectorAll('[data-until]').forEach(function(el){
    if(el.getAttribute('data-until')<today) el.remove();
  });
  document.querySelectorAll('[data-festblock]').forEach(function(b){
    if(!b.querySelector('[data-until]')) b.remove();
  });
})();
</script>"""

def fest_rows_html(items):
    out = []
    for f in items:
        meta = []
        if f["time"]: meta.append(e(f["time"]))
        if f["temple"]: meta.append(e(f["temple"]))
        out.append('<li class="fest-row" data-until="%s">'
                   '<span class="fest-when">%s</span>'
                   '<span class="fest-what"><b>%s</b>%s%s</span></li>' % (
            e(f["end"]), e(pretty_date(f["date"], f["end"])), e(f["fest"]),
            ('<span class="fest-meta">%s</span>' % " \u00b7 ".join(meta)) if meta else "",
            ('<span class="fest-note">%s</span>' % e(f["details"])) if f["details"] else ""))
    return "".join(out)

def temple_fest_html(t):
    items = FEST_BY_TEMPLE.get((t["name"].lower(), t["area"].lower()), [])
    if not items: return ""
    return ('<div class="t-fest" data-festblock><p class="t-fest-h">Upcoming at this temple</p>'
            '<ul class="fest-list">%s</ul></div>' % fest_rows_html(items))

# ---------------------------------------------------------------- index
def build_index():
    faq_html, faq_ld = [], []
    for q, a in FAQS:
        a = sub(a, N=str(N), NLON=str(N_LON), NLEI=str(N_LEI), MAILTO=MAILTO)
        faq_html.append("<details><summary>%s</summary><p>%s</p></details>" % (e(q), a))
        import re as _re
        faq_ld.append({"@type": "Question", "name": q,
            "acceptedAnswer": {"@type": "Answer", "text": _re.sub(r"<[^>]+>", "", a)}})

    city_cards = "".join(
        '<a class="city-card" href="%s.html"><b>%s</b><span>%d temple%s</span></a>' %
        (slug, e(nm), len(BY_CITY[slug]), "" if len(BY_CITY[slug]) == 1 else "s")
        for nm, slug, la, lo, rad, z, bl in CITIES if BY_CITY[slug])

    feats = []
    for name, area, cslug, why in FEATURED:
        t = next((x for x in T if x["name"] == name and x["area"] == area), None)
        if not t: continue
        feats.append('<div class="feat-card"><h3>%s</h3><p class="where">%s · %s</p><p>%s</p>'
                     '<a href="%s.html#t-%d">View in %s →</a></div>' %
                     (e(t["name"]), e(t["area"]), e(t["region"]), e(why),
                      cslug, t["id"], e(dict((c[1], c[0]) for c in CITIES)[cslug])))

    if FEST:
        fest_section = (
          '<section class="section wrap" id="festivals" data-festblock>'
          '<div class="section-head"><h2>Upcoming Hindu festival dates</h2>'
          '<p class="section-sub">Dates for the year ahead. <b>UK temples do not all observe on the same day</b> '
          '— panchang and regional reckonings differ — so check with your temple before travelling. '
          'Confirmed temple dates appear on each temple’s listing as we receive them.</p></div>'
          '<ul class="fest-list big">%s</ul>'
          '<p class="fest-foot">Know your temple’s dates? <a href="%s">Send them in</a> and we’ll add them.</p>'
          '</section>' % (fest_rows_html(FEST_NATIONAL), MAILTO))
    else:
        fest_section = ""

    page = head(
        "UK Hindu Temples — Interactive Map & Directory of %d Mandirs" % N,
        "Find your nearest Hindu temple by postcode. An interactive map and complete directory of %d verified mandirs across England, Scotland, Wales and Northern Ireland — with addresses, traditions, opening hours and directions." % N,
        BASE + "/", active="home", leaflet=True,
        extra="\n".join([
            ld({"@context": "https://schema.org", "@type": "WebSite", "name": "Hindu Temples UK",
                "url": BASE + "/", "description": "Interactive map and directory of Hindu temples across the United Kingdom.",
                "potentialAction": {"@type": "SearchAction",
                    "target": {"@type": "EntryPoint", "urlTemplate": BASE + "/?postcode={postcode}"},
                    "query-input": "required name=postcode"}}),
            ld({"@context": "https://schema.org", "@type": "Organization", "name": "Hindu Temples UK",
                "url": BASE + "/", "logo": BASE + "/favicon.svg", "email": EMAIL}),
            ld({"@context": "https://schema.org", "@type": "FAQPage", "mainEntity": faq_ld}),
        ]))
    page += header_html("home")
    page += sub("""
<section class="hero wrap">
  <p class="eyebrow">Sacred geography · United Kingdom</p>
  <h1>Find your nearest Hindu temple</h1>
  <p class="lede">From the hand-carved marble of Neasden to the Tamil kovils of Croydon and the ISKCON estate at Bhaktivedanta Manor — this directory maps <b>__N__ verified Hindu temples and mandirs</b> across England, Scotland, Wales and Northern Ireland. Search by postcode, or browse by region and tradition.</p>
  <ul class="hero-stats">
    <li><b>__N__</b>temples verified</li>
    <li><b>12</b>regions covered</li>
    <li><b>5</b>traditions mapped</li>
  </ul>
</section>

<section class="finder wrap" aria-label="Temple finder">
  <div class="search-card">
    <div class="search-row">
      <div class="field">
        <svg width="16" height="16" viewBox="0 0 24 24" fill="none" aria-hidden="true"><path d="M12 21s-7-6.4-7-11a7 7 0 1114 0c0 4.6-7 11-7 11z" stroke="#BE3524" stroke-width="1.6"/><circle cx="12" cy="10" r="2.4" stroke="#BE3524" stroke-width="1.6"/></svg>
        <input id="pc" type="text" inputmode="text" autocomplete="postal-code" placeholder="Enter a postcode (e.g. LE1 6EJ)" aria-label="Postcode" spellcheck="false"/>
      </div>
      <button class="btn btn-primary" id="go">Find</button>
    </div>
    <div class="search-aux">
      <button class="btn btn-ghost btn-sm" id="geo">◎ Use my location</button>
      <button class="btn btn-ghost btn-sm" id="pin" aria-pressed="false">✛ Drop a pin</button>
    </div>
    <p class="hint" id="hint" role="status" aria-live="polite">Nearest temples are ranked by straight-line distance from your postcode.</p>
    <div class="filters">
      <div class="sel"><label for="fRegion">Region</label>
        <select id="fRegion"><option value="">All regions</option></select></div>
      <div class="sel"><label for="fCounty">County / borough</label>
        <select id="fCounty"><option value="">All counties</option></select></div>
      <div class="sel"><label for="fTrad">Tradition</label>
        <select id="fTrad"><option value="">All traditions</option></select></div>
      <div class="sel"><label for="fName">Search by name</label>
        <div class="field" style="min-height:42px"><input id="fName" type="search" placeholder="Temple name…" style="font-family:var(--font-body);font-size:15px;padding:9px 0"/></div></div>
    </div>
    <p class="report-line">Missing temple or wrong details? <a href="__MAILTO__">Report it →</a></p>
  </div>

  <div class="map-wrap">
    <div id="map" role="application" aria-label="Map of Hindu temples in the UK"></div>
    <p class="map-status" id="mapStatus">Loading __N__ temples…</p>
    <div class="legend collapsed" id="legend">
      <button class="legend-h" id="legendToggle" aria-expanded="false" aria-controls="legendRows">Key <span class="caret">▾</span></button>
      <div class="legend-body" id="legendRows"></div>
    </div>
  </div>

  <div class="results-block">
    <div class="results-head">
      <span class="count" id="count" role="status" aria-live="polite"></span>
      <button class="linklike" id="clear">Clear filters &amp; location</button>
    </div>
    <div class="results" id="results" aria-label="Temple results">
      <div class="card skel" aria-hidden="true"></div>
      <div class="card skel" aria-hidden="true"></div>
      <div class="card skel" aria-hidden="true"></div>
    </div>
  </div>
</section>

__FESTSECTION__
<section class="section wrap" id="cities">
  <div class="section-head">
    <h2>Browse temples by city</h2>
    <p class="section-sub">Every major Hindu community in Britain, from the capital to the Golden Mile — each city page lists its mandirs with addresses, traditions and directions.</p>
  </div>
  <div class="city-grid">__CITYCARDS__</div>
</section>

<section class="section wrap" id="featured">
  <div class="section-head">
    <h2>Landmark temples of the UK</h2>
    <p class="section-sub">Six places that anchor Hindu life in Britain.</p>
  </div>
  <div class="feat-grid">__FEATS__</div>
</section>

<section class="section wrap" id="faq">
  <div class="section-head">
    <h2>Common questions</h2>
  </div>
  <div class="faq">__FAQS__</div>
</section>
""", N=str(N), CITYCARDS=city_cards, FEATS="".join(feats), FAQS="".join(faq_html),
     MAILTO=MAILTO, FESTSECTION=fest_section)
    page += footer_html()
    page += """
<script src="https://cdnjs.cloudflare.com/ajax/libs/leaflet/1.9.4/leaflet.min.js" defer></script>
<script src="app.js" defer></script>"""
    open("index.html", "w", encoding="utf-8").write(page)

# ---------------------------------------------------------------- city pages
def build_city_pages():
    made = []
    for nm, slug, la, lo, rad, z, blurb in CITIES:
        temples = BY_CITY[slug]
        if not temples: continue
        n = len(temples)
        clat = sum(t["lat"] for t in temples) / n
        clng = sum(t["lng"] for t in temples) / n
        examples = ", ".join(t["name"] for t in temples[:3])
        title = "Hindu Temples in %s — %d Mandir%s Listed" % (nm, n, "" if n == 1 else "s")
        desc = ("Find Hindu temples and mandirs in %s: addresses, opening hours, traditions and directions for %d temple%s including %s." % (
            nm, n, "" if n == 1 else "s", examples))[:158]
        canon = "%s/%s.html" % (BASE, slug)
        body = sub("""
<div class="page wrap">
  <nav class="crumb" aria-label="Breadcrumb"><a href="index.html">Home</a> › <a href="directory.html">Directory</a> › __NM__</nav>
  <h1>Hindu Temples in __NM__</h1>
  <div class="prose">
    <p class="lede">__BLURB__</p>
    <p><b>__CN__ temple__PL__</b> listed in and around __NM__. <a href="index.html?lat=__LAT__&amp;lng=__LNG__&amp;z=__Z__">Open __NM__ on the interactive map</a> to search by postcode and see them all at once.</p>
  </div>
  <div class="tlist">__TEMPLES__</div>
  <p class="backlink"><a href="directory.html">See every UK Hindu temple by region →</a></p>
</div>""", NM=e(nm), BLURB=e(blurb), CN=str(n), PL="" if n == 1 else "s",
           LAT="%.4f" % clat, LNG="%.4f" % clng, Z=str(z),
           TEMPLES="".join(temple_article(t) for t in temples))
        page = (head(title, desc, canon, leaflet=False,
                     extra="\n".join([ld(ld_itemlist(temples, "Hindu Temples in %s" % nm)),
                                      ld(ld_breadcrumb([("Home",""),("Directory","directory.html"),(nm, slug+".html")]))]))
                + header_html("") + body + footer_html())
        open("%s.html" % slug, "w", encoding="utf-8").write(page)
        made.append((nm, slug, n))
    return made

# ---------------------------------------------------------------- directory
def build_directory():
    jump = " ".join('<a href="#%s">%s</a>' % (r.lower().replace(" ", "-"), e(r))
                    for r in REGIONS if REGION_COUNT[r])
    secs = []
    for r in REGIONS:
        ts = sorted((t for t in T if t["region"] == r), key=lambda x: (x["county"], x["name"]))
        if not ts: continue
        rows = []
        for t in ts:
            loc = " · ".join(x for x in [t["area"], t["county"]] if x)
            if t.get("_city"):
                link, rel = "%s.html#t-%d" % (t["_city"], t["id"]), ""
            else:
                link, rel = gdir(t), ' target="_blank" rel="noopener"'
            closed = ' <span class="tag closed">Closed</span>' if str(t.get("status","")).lower()=="closed" else ""
            site = ('  <a class="dsite" href="%s" target="_blank" rel="noopener">Website ↗</a>'
                    % e(t["url"])) if t.get("url") else ""
            rows.append('<li><a href="%s"%s>%s</a> <span class="dloc">%s</span>%s%s</li>' %
                        (link, rel, e(t["name"]), e(loc), closed, site))
        secs.append('<section class="region-sec"><h2 id="%s">%s <span class="rn">%d temple%s</span></h2>'
                    '<ul class="dir">%s</ul></section>' %
                    (r.lower().replace(" ", "-"), e(r), len(ts), "" if len(ts)==1 else "s", "".join(rows)))
    body = sub("""
<div class="page wrap">
  <nav class="crumb" aria-label="Breadcrumb"><a href="index.html">Home</a> › Directory</nav>
  <h1>UK Hindu Temple Directory</h1>
  <div class="prose">
    <p class="lede">Every Hindu temple and mandir in this directory — <b>__N__ in total</b> — listed by region with its town and county. Prefer a map? Use the <a href="index.html">postcode finder</a> or the <a href="temple-map.html">all-UK temple map</a>.</p>
  </div>
  <div class="dir-tools">
    <div class="field">
      <svg width="15" height="15" viewBox="0 0 24 24" fill="none" aria-hidden="true"><circle cx="11" cy="11" r="7" stroke="#5F5344" stroke-width="1.6"/><path d="M20 20l-4-4" stroke="#5F5344" stroke-width="1.6"/></svg>
      <input id="dirFilter" type="search" placeholder="Filter by name or town…" aria-label="Filter temples"/>
    </div>
    <span class="count" id="dirCount"></span>
  </div>
  <p class="jumper">Jump to: __JUMP__</p>
  __SECS__
</div>
<script>
(function(){
  var input=document.getElementById('dirFilter'),items=[].slice.call(document.querySelectorAll('ul.dir li')),
      secs=[].slice.call(document.querySelectorAll('.region-sec')),count=document.getElementById('dirCount');
  function apply(){
    var q=input.value.trim().toLowerCase(),shown=0;
    items.forEach(function(li){var on=!q||li.textContent.toLowerCase().indexOf(q)>-1;li.style.display=on?'':'none';if(on)shown++;});
    secs.forEach(function(s){var any=[].slice.call(s.querySelectorAll('li')).some(function(li){return li.style.display!=='none';});s.style.display=any?'':'none';});
    count.textContent=q?shown+' of __N__ temples':'';
  }
  input.addEventListener('input',apply);
})();
</script>""", N=str(N), JUMP=jump, SECS="".join(secs))
    page = (head("UK Hindu Temple Directory — All %d Mandirs by Region" % N,
                 "The complete A–Z directory of %d Hindu temples and mandirs across the UK, organised by region and county, with links to full details, maps and directions." % N,
                 BASE + "/directory.html", active="directory",
                 extra="\n".join([ld(ld_itemlist(T, "UK Hindu Temple Directory", compact=True)),
                                  ld(ld_breadcrumb([("Home",""),("Directory","directory.html")]))]))
            + header_html("directory") + body + footer_html())
    open("directory.html", "w", encoding="utf-8").write(page)

# ---------------------------------------------------------------- temple map
def build_temple_map():
    rows = "".join("<tr><td>%s</td><td>%d</td></tr>" % (e(r), REGION_COUNT[r])
                   for r in REGIONS if REGION_COUNT[r])
    trad_counts = {}
    for t in T: trad_counts[t["trad"]] = trad_counts.get(t["trad"], 0) + 1
    trad_line = ", ".join("%d %s" % (trad_counts[k], TRAD_LABEL[k].lower())
                          for k in ["General","Saiva","BAPS","Swaminarayan","ISKCON"] if trad_counts.get(k))
    pts = [{"lat": t["lat"], "lng": t["lng"], "n": t["name"], "a": t["area"], "h": bool(t.get("hl"))} for t in T]
    body = sub("""
<div class="page wrap">
  <nav class="crumb" aria-label="Breadcrumb"><a href="index.html">Home</a> › Temple map</nav>
  <h1>Every Hindu temple in the UK, on one map</h1>
  <div class="prose">
    <p class="lede">All <b>__N__ Hindu temples and mandirs</b> in this directory, plotted across England, Scotland, Wales and Northern Ireland — from Aberdeen to Southampton, and from Belfast to Ipswich. Each dot is one temple; tap it for the name.</p>
  </div>
  <div class="bigmap-wrap"><div id="bigmap" role="application" aria-label="Map showing every Hindu temple in the UK"></div></div>
  <p class="map-cap">__N__ temples · one dot each · featured temples in vermilion · source: hindu-temples.uk</p>
  <div class="prose">
    <h2>Hindu temples by region</h2>
    <p>London and the Midlands hold the largest clusters, tracing where Britain's Hindu communities first settled — but every UK region now has at least one mandir. The directory spans __TRADS__.</p>
  </div>
  <table class="stats-table">
    <thead><tr><th scope="col">Region</th><th scope="col">Temples</th></tr></thead>
    <tbody>__ROWS__<tr><td><b>United Kingdom</b></td><td><b>__N__</b></td></tr></tbody>
  </table>
  <div class="prose">
    <p>Explore the same data with postcode search and filters on the <a href="index.html">interactive finder</a>, browse the <a href="directory.html">full A–Z directory</a>, or jump to a city such as <a href="london.html">London</a>, <a href="leicester.html">Leicester</a> or <a href="birmingham.html">Birmingham</a>.</p>
  </div>
</div>
<script src="https://cdnjs.cloudflare.com/ajax/libs/leaflet/1.9.4/leaflet.min.js" defer></script>
<script>
document.addEventListener('DOMContentLoaded',function(){
  var P=__PTS__;
  var m=L.map('bigmap',{scrollWheelZoom:false}).setView([54.5,-3.4],5);
  L.tileLayer('https://tile.openstreetmap.org/{z}/{x}/{y}.png',
    {maxZoom:12,attribution:'&copy; <a href="https://www.openstreetmap.org/copyright">OpenStreetMap</a> contributors'}).addTo(m);
  P.forEach(function(p){
    L.circleMarker([p.lat,p.lng],{radius:p.h?6:4,color:'#fff',weight:1,
      fillColor:p.h?'#BE3524':'#B8922F',fillOpacity:.88})
      .bindTooltip(p.n+(p.a?' — '+p.a:''),{direction:'top'}).addTo(m);
  });
});
</script>""", N=str(N), ROWS=rows, TRADS=e(trad_line),
           PTS=json.dumps(pts, ensure_ascii=False, separators=(",", ":")))
    page = (head("Map of Every Hindu Temple in the UK — %d Mandirs Plotted" % N,
                 "A single map of all %d Hindu temples and mandirs in the United Kingdom, with counts by region. See where Britain's Hindu communities worship, from London to Glasgow." % N,
                 BASE + "/temple-map.html", active="map", leaflet=True,
                 extra="\n".join([
                     ld({"@context": "https://schema.org", "@type": "Dataset",
                         "name": "Hindu temples of the United Kingdom",
                         "description": "Locations of %d verified Hindu temples and mandirs across the UK." % N,
                         "url": BASE + "/temple-map.html", "license": BASE + "/about.html",
                         "spatialCoverage": "United Kingdom", "creator": {"@type": "Organization", "name": "Hindu Temples UK"}}),
                     ld(ld_breadcrumb([("Home",""),("Temple map","temple-map.html")]))]))
            + header_html("map") + body + footer_html())
    open("temple-map.html", "w", encoding="utf-8").write(page)

# ---------------------------------------------------------------- about
def build_about():
    body = sub("""
<div class="page wrap">
  <nav class="crumb" aria-label="Breadcrumb"><a href="index.html">Home</a> › About</nav>
  <h1>About this directory</h1>
  <div class="prose">
    <p class="lede">hindu-temples.uk is a free, independent directory of Hindu temples and mandirs across the United Kingdom — built because nothing like it existed. Directories of other places of worship are widely published; the UK's __N__-plus mandirs deserved the same visibility.</p>

    <h2>How the data is compiled</h2>
    <p>The directory brings together the national temple bodies' published listings, community directories and multiple independent research passes, then verifies <b>every single temple</b> against live mapping data. Verification confirms three things: the temple exists at a real street address, its map pin sits on the actual building rather than a town centre, and it is genuinely a Hindu place of worship. Entries that fail any check are removed.</p>
    <p>Each listing records the temple's name, address, town, county, region, tradition (Sanatan/community, Saiva/Tamil, BAPS, other Swaminarayan, or ISKCON), and — where verified — phone number, opening hours and website. Currently: <b>__N__ temples</b> across all 12 UK regions, last updated <b>__TODAY__</b>.</p>

    <h2>What the directory is not</h2>
    <p>It is not exhaustive — smaller community shrines and home gatherings open, move and close all the time — and it is not affiliated with any temple, organisation or tradition. Distances on the finder are straight-line, for ranking; use each temple's Directions link for an actual route. Always confirm opening times with the temple before travelling, especially around festivals.</p>

    <h2 id="report">Report a temple, closure or correction</h2>
    <p>Spotted a missing mandir, a closure, wrong details, or have opening hours or a website to add? Email <a href="__MAILTO__">__EMAIL__</a> with the temple's name and address and what needs changing. Every submission is checked before the directory is updated.</p>

    <h2>Use of this data</h2>
    <p>The directory is free for personal and community use — linking to <a href="index.html">hindu-temples.uk</a> is appreciated. For bulk or commercial use, please get in touch first.</p>
  </div>
</div>""", N=str(N), TODAY=TODAY, MAILTO=MAILTO, EMAIL=EMAIL)
    page = (head("About Hindu Temples UK — Data, Sources & Contact",
                 "How hindu-temples.uk compiles and verifies the UK's most complete directory of Hindu temples — and how to report a missing temple, a closure or a correction.",
                 BASE + "/about.html", active="about",
                 extra="\n".join([
                     ld({"@context": "https://schema.org", "@type": "AboutPage",
                         "name": "About Hindu Temples UK", "url": BASE + "/about.html"}),
                     ld({"@context": "https://schema.org", "@type": "Organization",
                         "name": "Hindu Temples UK", "url": BASE + "/", "email": EMAIL,
                         "contactPoint": {"@type": "ContactPoint", "email": EMAIL, "contactType": "customer support"}}),
                     ld(ld_breadcrumb([("Home",""),("About","about.html")]))]))
            + header_html("about") + body + footer_html())
    open("about.html", "w", encoding="utf-8").write(page)

# ---------------------------------------------------------------- 404 + legacy redirect
def build_404():
    body = """
<div class="page wrap">
  <h1>That page doesn't exist</h1>
  <div class="prose">
    <p class="lede">The address may have changed, or the link was mistyped. Everything on the site is one step away:</p>
    <p><a class="btn btn-primary" href="index.html" style="margin-right:10px">Find a temple</a>
       <a class="btn btn-ghost" href="directory.html">Browse the full directory</a></p>
  </div>
</div>"""
    page = (head("Page not found — Hindu Temples UK", "This page could not be found.",
                 BASE + "/404.html", noindex=True) + header_html("") + body + footer_html())
    open("404.html", "w", encoding="utf-8").write(page)

def build_legacy_redirect():
    # map-comparison.html was briefly live; forward it cleanly to temple-map.html
    open("map-comparison.html", "w", encoding="utf-8").write(
"""<!DOCTYPE html>
<html lang="en-GB">
<head>
<meta charset="utf-8"/>
<title>Moved — Map of Every Hindu Temple in the UK</title>
<meta name="robots" content="noindex"/>
<link rel="canonical" href="%s/temple-map.html"/>
<meta http-equiv="refresh" content="0; url=temple-map.html"/>
</head>
<body>
<p>This page has moved to <a href="temple-map.html">the UK Hindu temple map</a>.</p>
</body>
</html>""" % BASE)

# ---------------------------------------------------------------- Hinduism in Britain
TIMELINE = [
 ("1929", "Shanti Sadan opens in London",
  "Hari Prasad Shastri, a teacher of Adhyatma Yoga, settles in London and founds Shanti Sadan \u2014 among the earliest organised Hindu institutions in Britain, decades before mass migration.", ""),
 ("1950s", "The first congregations",
  "Hindus begin arriving in numbers after the Second World War, mostly from Gujarat and Punjab. Worship happens in front rooms and hired halls; there are no temples yet.", ""),
 ("1960s", "ISKCON reaches London",
  "Devotees sent from San Francisco establish a base in a Covent Garden warehouse. George Harrison takes an interest, records with them at Abbey Road, and co-signs the lease on their Bloomsbury premises.", ""),
 ("1969", "Leicester: murtis installed on Cromford Street",
  "On 20 July 1969, Radha-Krishna murtis are installed at a converted property in Highfields, Leicester \u2014 widely cited as the first Hindu temple in Britain. A building becomes a mandir only once the murtis are installed.", "leicester"),
 ("1970", "BAPS opens in a disused Islington church",
  "The Swaminarayan community converts a redundant church in Islington \u2014 typical of the era, when almost every early mandir was a repurposed chapel, hall or shop.", ""),
 ("1972", "The Ugandan expulsion",
  "Idi Amin expels Uganda's Asian population. Around 60,000 people are forced out; the British government admits some 27,000 through the Uganda Resettlement Board. Many are Gujarati Hindus who arrive having lost homes and businesses. Leicester, Wembley and the East Midlands are transformed.", ""),
 ("1973", "George Harrison gives Bhaktivedanta Manor",
  "Harrison buys the Hertfordshire estate and donates it to ISKCON. It becomes the movement's UK home and today hosts the largest Janmashtami celebration outside India.", "watford"),
 ("1980", "Purpose-run mandirs take shape",
  "Shri Nathji Sanatan Hindu Mandir opens in Leytonstone under Shri Vallabh Nidhi UK, as communities move from borrowed rooms to buildings of their own.", ""),
 ("1981", "Wimbledon: the first fully consecrated temple",
  "Shree Ghanapathy Temple opens in a former church hall in Wimbledon, redesigned to Hindu architectural guidelines \u2014 described as the first fully consecrated Hindu temple in Europe.", "london"),
 ("1982", "Neasden begins",
  "Having outgrown Islington, BAPS moves to a converted warehouse in Neasden \u2014 the site that will become Britain's best-known mandir.", ""),
 ("1995", "Europe's first traditional stone mandir",
  "On 20 August 1995, BAPS Shri Swaminarayan Mandir opens in Neasden: hand-carved marble and limestone, built by traditional methods at a cost of around \u00a312 million. The first purpose-built stone mandir in Europe, as distinct from a converted secular building.", "london"),
 ("2010", "A second generation of building",
  "Shree Sanatan Hindu Mandir is completed off Ealing Road, Wembley \u2014 part of a wave of purpose-built mandirs replacing the converted halls of the 1970s.", "london"),
 ("2021", "Over a million",
  "The census records 1,066,894 Hindus across the UK \u2014 the third-largest religious group, and more than triple the 1981 figure.", ""),
]

POP = [("1961",30000),("1971",138000),("1981",278000),("1991",397000),
       ("2001",558810),("2011",835394),("2021",1066894)]

REGION_POP = [("Greater London","453,034","5.1%"),("South East","154,748","1.7%"),
              ("East Midlands","120,345","2.5%"),("West Midlands","88,116","1.5%")]

NATIONS = [("England","1,020,533","1.8%","2021"),("Scotland","29,929","0.6%","2022"),
           ("Wales","12,242","0.4%","2021"),("Northern Ireland","4,190","0.2%","2021")]

def pop_chart_svg():
    W, H, PAD_L, PAD_B, PAD_T = 700, 300, 52, 40, 16
    mx = max(v for _, v in POP)
    bw = (W - PAD_L - 16) / len(POP)
    bars = []
    for i, (yr, v) in enumerate(POP):
        h = (H - PAD_B - PAD_T) * v / mx
        x = PAD_L + i * bw + bw * 0.18
        y = H - PAD_B - h
        w = bw * 0.64
        lab = ("%.2fm" % (v / 1e6)) if v >= 1e6 else ("%dk" % round(v / 1000))
        bars.append(
          '<rect x="%.1f" y="%.1f" width="%.1f" height="%.1f" rx="3" fill="url(#bg1)"/>'
          '<text x="%.1f" y="%.1f" text-anchor="middle" class="v">%s</text>'
          '<text x="%.1f" y="%d" text-anchor="middle" class="x">%s</text>'
          % (x, y, w, h, x + w/2, y - 6, lab, x + w/2, H - PAD_B + 20, yr))
    grid = []
    for f in (0, .25, .5, .75, 1):
        gy = H - PAD_B - (H - PAD_B - PAD_T) * f
        grid.append('<line x1="%d" y1="%.1f" x2="%d" y2="%.1f" class="g"/>'
                    '<text x="%d" y="%.1f" text-anchor="end" class="y">%s</text>'
                    % (PAD_L, gy, W - 8, gy, PAD_L - 8, gy + 4,
                       "0" if f == 0 else "%.2fm" % (mx * f / 1e6)))
    return ('<svg viewBox="0 0 %d %d" role="img" class="popchart" '
            'aria-label="Hindu population of the UK by census year, rising from 30,000 in 1961 to 1,066,894 in 2021">'
            '<defs><linearGradient id="bg1" x1="0" y1="0" x2="0" y2="1">'
            '<stop offset="0%%" stop-color="#E8A13A"/><stop offset="100%%" stop-color="#BE3524"/>'
            '</linearGradient></defs>%s%s</svg>' % (W, H, "".join(grid), "".join(bars)))

def build_history():
    city_slugs = {c[1] for c in CITIES}
    items = []
    for yr, tl_title, tl_text, link in TIMELINE:
        more = ('  <a href="%s.html">See temples in %s \u2192</a>'
                % (link, dict((c[1], c[0]) for c in CITIES)[link])) if link in city_slugs else ""
        items.append('<li class="tl-item"><span class="tl-dot" aria-hidden="true"></span>'
                     '<span class="tl-year">%s</span>'
                     '<div class="tl-body"><h3>%s</h3><p>%s%s</p></div></li>'
                     % (e(yr), e(tl_title), e(tl_text), more))
    nat = "".join("<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>"
                  % (e(a), e(b), e(c), e(d)) for a, b, c, d in NATIONS)
    reg = "".join("<tr><td>%s</td><td>%s</td><td>%s</td></tr>"
                  % (e(a), e(b), e(c)) for a, b, c in REGION_POP)

    body = sub("""
<div class="page wrap">
  <nav class="crumb" aria-label="Breadcrumb"><a href="index.html">Home</a> › Hinduism in Britain</nav>
  <h1>Hinduism in Britain: a short history</h1>
  <div class="prose">
    <p class="lede">Britain's Hindu community grew from roughly <b>30,000 people in 1961</b> to <b>over a million by 2021</b> — and the __N__ mandirs in this directory are the physical record of that. Most began as converted church halls, shops and front rooms. This is how they came to be here.</p>

    <h2>Before the migration</h2>
    <p>Hindu presence in Britain long predates the post-war era, but organised institutions were rare. In <b>1929</b> the scholar Hari Prasad Shastri founded <b>Shanti Sadan</b> in London, one of the earliest lasting Hindu centres in the country. For decades afterwards, worship for the small resident community happened privately.</p>

    <h2>Three waves</h2>
    <p>Hindu migration to Britain came in three broad waves. The first, after the Second World War, brought people directly from <b>Gujarat and Punjab</b> to fill labour shortages in industrial towns. The second, and the most consequential, came from <b>East Africa</b> in the late 1960s and early 1970s. The third, more recent, has drawn Hindus from across India and from the wider diaspora — Sri Lanka, Mauritius, Fiji, Guyana, Trinidad — which is why the traditions in this directory range from Gujarati Swaminarayan to Tamil Saiva.</p>

    <h2>1972: the year that changed the map</h2>
    <p>In August 1972, President Idi Amin ordered the expulsion of Uganda's Asian population. Around <b>60,000 people</b> were forced to leave, and the British government admitted some <b>27,000</b> through the Uganda Resettlement Board. Many were Gujarati Hindus who arrived with almost nothing, having lost homes and businesses built over generations.</p>
    <p>They settled where there was work and where community already existed — <b>Leicester</b>, <b>Wembley and Brent</b>, <b>Coventry</b>, <b>Bolton</b>, <b>Preston</b>. That single event explains much of the map on this site: it is why Leicester's Belgrave Road became the Golden Mile, and why <a href="london.html">London</a> and the <a href="leicester.html">East Midlands</a> hold the densest clusters of mandirs in the country today.</p>

    <h2>From front rooms to mandirs</h2>
    <p>The earliest temples were not built as temples. They were disused churches, chapels, warehouses, shops and terraced houses, bought cheaply by communities with little money and consecrated by installing the murtis — because a building becomes a mandir only when the deities are installed in it. That pattern is still visible: many temples in this directory occupy buildings that began life as something else.</p>
  </div>

  <h2 class="tl-h">Timeline</h2>
  <ol class="tl">__ITEMS__</ol>

  <div class="prose">
    <h2>The numbers</h2>
    <p>Religious affiliation was not recorded in the census before 2001, so earlier figures are estimates. The trajectory is nonetheless clear.</p>
  </div>
  __CHART__
  <p class="map-cap">Hindu population of the United Kingdom by year. Pre-2001 figures are estimates.</p>

  <div class="prose"><h2>Where Hindus live in Britain today</h2></div>
  <table class="stats-table">
    <thead><tr><th scope="col">Nation</th><th scope="col">Hindu population</th><th scope="col">Share</th><th scope="col">Census</th></tr></thead>
    <tbody>__NAT__</tbody>
  </table>
  <table class="stats-table" style="margin-top:14px">
    <thead><tr><th scope="col">Region (England)</th><th scope="col">Hindu population</th><th scope="col">Share</th></tr></thead>
    <tbody>__REG__</tbody>
  </table>

  <div class="prose">
    <p style="margin-top:22px">Greater London alone accounts for over 450,000 Hindus — around <b>5.1%</b> of its population, the highest concentration in the UK. You can see the same pattern in this directory: <a href="london.html">London</a> holds more mandirs than any other city, followed by <a href="leicester.html">Leicester</a>.</p>

    <h2>A note on "the first Hindu temple in Britain"</h2>
    <p>Several temples are described as the first, and they are describing different things. The Leicester mandir of <b>1969</b> is generally cited as the earliest Hindu temple in Britain, in a converted property. <a href="london.html">Shree Ghanapathy Temple</a> in Wimbledon (<b>1981</b>) is described as the first fully consecrated Hindu temple in Europe. <a href="london.html">BAPS Shri Swaminarayan Mandir</a> at Neasden (<b>1995</b>) was Europe's first traditional stone mandir, purpose-built rather than converted. All three claims can be true at once.</p>

    <h2>Sources</h2>
    <p class="srcs">Population figures: Office for National Statistics, <b>Census 2021</b> (England and Wales); National Records of Scotland, <b>Census 2022</b>; Northern Ireland Statistics and Research Agency, <b>Census 2021</b>. Pre-2001 estimates and migration history via published summaries of UK census and migration records. Ugandan expulsion figures: Uganda Resettlement Board records as summarised in standard histories of the 1972 expulsion. Temple dates: each temple's own published history, and contemporary reporting including the <i>Leicester Mercury</i>, 21 July 1969.</p>
    <p class="srcs">Where sources disagree — particularly on which temple was "first" — this page says so rather than choosing. Spotted an error or have a documented date? <a href="__MAILTO__">Tell us</a> and we'll correct it.</p>
  </div>

  <p class="backlink"><a href="directory.html">Browse all __N__ UK Hindu temples →</a></p>
</div>""", N=str(N), ITEMS="".join(items), CHART=pop_chart_svg(),
           NAT=nat, REG=reg, MAILTO=MAILTO)

    page = (head("Hinduism in Britain: A Short History of UK Hindu Temples",
        "How Britain's Hindu community grew from 30,000 in 1961 to over a million — the 1972 Ugandan expulsion, the first mandirs, and the temples that record that history.",
        BASE + "/hinduism-in-britain.html", active="history",
        extra="\n".join([
            ld({"@context": "https://schema.org", "@type": "Article",
                "headline": "Hinduism in Britain: a short history",
                "description": "The history of Hindu migration to Britain and the temples it produced, from 1929 to the 2021 census.",
                "author": {"@type": "Organization", "name": "Hindu Temples UK"},
                "publisher": {"@type": "Organization", "name": "Hindu Temples UK",
                              "logo": {"@type": "ImageObject", "url": BASE + "/favicon.svg"}},
                "datePublished": TODAY, "dateModified": TODAY,
                "mainEntityOfPage": BASE + "/hinduism-in-britain.html"}),
            ld(ld_breadcrumb([("Home", ""), ("Hinduism in Britain", "hinduism-in-britain.html")]))]))
        + header_html("history") + body + footer_html())
    open("hinduism-in-britain.html", "w", encoding="utf-8").write(page)

# ---------------------------------------------------------------- learn pages
TRADITIONS = [
 ("Saiva", "Saiva / Tamil kovils",
  "Temples in the South Indian and Sri Lankan Tamil tradition, usually dedicated to Shiva, Murugan, Ganesha (Pillaiyar) or a form of the Goddess (Amman). Britain's Tamil kovils largely date from the 1980s onward, established by Tamil communities from Sri Lanka and South India.",
  "A tall stepped gateway tower (gopuram) where the building allows it, a metal flagstaff (kodimaram) facing the main shrine, camphor flame and bell at puja, and worship conducted in Tamil and Sanskrit. Priests are often bare-chested with the sacred thread.",
  ["Shree Ghanapathy Temple, Wimbledon", "Highgate Hill Murugan Temple, London", "London Sri Murugan Temple, Manor Park"]),
 ("BAPS", "BAPS Swaminarayan mandirs",
  "Part of the Bochasanwasi Shri Akshar Purushottam Swaminarayan Sanstha, a worldwide Gujarati organisation. BAPS built Britain's most recognisable mandir at Neasden in 1995 and runs a highly organised network of centres with youth and community programmes.",
  "Carved marble and limestone in traditional North Indian style, immaculate order, shoes and belongings left in dedicated areas, separate seating for men and women, and structured youth activity. Neasden also has a permanent 'Understanding Hinduism' exhibition.",
  ["BAPS Shri Swaminarayan Mandir, Neasden", "BAPS Shri Swaminarayan Mandir, Leicester", "BAPS Shri Swaminarayan Mandir, Wellingborough"]),
 ("ISKCON", "ISKCON / Vaishnava temples",
  "The International Society for Krishna Consciousness, in Britain since the 1960s. Its UK home, Bhaktivedanta Manor, was given by George Harrison in 1973 and hosts the largest Janmashtami celebration outside India.",
  "Continuous kirtan and drums, Radha-Krishna murtis, a strong emphasis on vegetarian prasad served to all visitors, and a mix of Indian and British-born devotees. Often the most welcoming to complete newcomers.",
  ["Bhaktivedanta Manor, Aldenham", "ISKCON London Radha-Krishna Temple, Soho", "ISKCON Leicester"]),
 ("Swaminarayan", "Other Swaminarayan mandirs",
  "Swaminarayan Hinduism has several distinct branches beyond BAPS \u2014 including the Kutch Satsang, SKS, ISSO and Swaminarayan Gadi traditions. Each has its own lineage and its own mandirs, mostly serving Gujarati communities.",
  "Similar devotional style to BAPS but organisationally separate, often in converted or purpose-built halls, with Gujarati as the main community language and strong satsang (congregational) activity.",
  ["Shree Swaminarayan Mandir, Kingsbury", "Shree Kutch Satsang Swaminarayan Temple, Kenton", "Swaminarayan Hindu Temple, Crawley"]),
 ("General", "Sanatan and community mandirs",
  "The largest group in this directory, and the hardest to summarise. These are community temples \u2014 often called Sanatan Mandir, Hindu Temple, or Hindu Cultural Centre \u2014 serving all Hindus in a town regardless of regional background. Many are the oldest temples in Britain.",
  "Multiple deities side by side rather than one focus, since the temple serves the whole local community. Frequently housed in a converted church, chapel or civic hall. These double as cultural centres, running language classes, weddings and festivals.",
  ["Shree Sanatan Mandir, Leicester", "Shri Sanatan Hindu Mandir, Wembley", "Bharat Hindu Samaj Mandir, Peterborough"]),
]

def svg_temple_types():
    return '''<svg viewBox="0 0 720 300" class="dgm" role="img" aria-label="Three types of Hindu temple building found in Britain: a North Indian shikhara tower, a South Indian gopuram gateway tower, and a converted church or hall.">
<g class="dgm-ink">
  <!-- 1. Nagara / shikhara -->
  <g transform="translate(20,0)">
    <path d="M100 40 C112 90 122 140 128 172 L72 172 C78 140 88 90 100 40 Z" class="f1"/>
    <path d="M100 40 C112 90 122 140 128 172 L72 172 C78 140 88 90 100 40 Z" class="s"/>
    <ellipse cx="100" cy="36" rx="15" ry="6" class="f2"/><ellipse cx="100" cy="36" rx="15" ry="6" class="s"/>
    <path d="M100 34 L100 22 M94 22 q6 -12 12 0 z" class="s"/>
    <circle cx="100" cy="16" r="4" class="f3"/>
    <rect x="52" y="172" width="96" height="52" rx="3" class="f2"/><rect x="52" y="172" width="96" height="52" rx="3" class="s"/>
    <rect x="88" y="192" width="24" height="32" rx="2" class="f1"/><rect x="88" y="192" width="24" height="32" rx="2" class="s"/>
    <line x1="40" y1="224" x2="160" y2="224" class="s"/>
    <text x="100" y="252" class="lbl">North Indian (Nagara)</text>
    <text x="100" y="270" class="sub">Curved shikhara tower</text>
    <text x="100" y="286" class="ex">e.g. Neasden</text>
  </g>
  <!-- 2. Dravidian / gopuram -->
  <g transform="translate(250,0)">
    <path d="M62 172 L74 60 L126 60 L138 172 Z" class="f1"/><path d="M62 172 L74 60 L126 60 L138 172 Z" class="s"/>
    <line x1="70" y1="112" x2="130" y2="112" class="s"/><line x1="66" y1="142" x2="134" y2="142" class="s"/>
    <line x1="74" y1="86" x2="126" y2="86" class="s"/>
    <path d="M70 60 q30 -16 60 0" class="f2"/><path d="M70 60 q30 -16 60 0" class="s"/>
    <circle cx="84" cy="44" r="3.5" class="f3"/><circle cx="100" cy="41" r="3.5" class="f3"/><circle cx="116" cy="44" r="3.5" class="f3"/>
    <rect x="86" y="126" width="28" height="46" rx="2" class="f2"/><rect x="86" y="126" width="28" height="46" rx="2" class="s"/>
    <line x1="30" y1="172" x2="30" y2="70" class="s2"/>
    <path d="M30 70 l0 -6 M25 64 h10" class="s"/><circle cx="30" cy="60" r="3" class="f3"/>
    <text x="30" y="190" class="tag2">kodimaram</text>
    <line x1="14" y1="172" x2="160" y2="172" class="s"/>
    <text x="100" y="252" class="lbl">South Indian (Dravidian)</text>
    <text x="100" y="270" class="sub">Stepped gopuram + flagstaff</text>
    <text x="100" y="286" class="ex">e.g. Balaji, Tividale</text>
  </g>
  <!-- 3. Converted building -->
  <g transform="translate(480,0)">
    <path d="M58 172 L58 106 L100 74 L142 106 L142 172 Z" class="f1"/>
    <path d="M58 172 L58 106 L100 74 L142 106 L142 172 Z" class="s"/>
    <path d="M100 74 L100 62 M94 62 q6 -11 12 0 z" class="s"/><circle cx="100" cy="56" r="3.5" class="f3"/>
    <rect x="88" y="128" width="24" height="44" rx="12" class="f2"/><rect x="88" y="128" width="24" height="44" rx="12" class="s"/>
    <path d="M70 118 a6 6 0 0 1 12 0 v14 h-12 z" class="f2"/><path d="M70 118 a6 6 0 0 1 12 0 v14 h-12 z" class="s"/>
    <path d="M118 118 a6 6 0 0 1 12 0 v14 h-12 z" class="f2"/><path d="M118 118 a6 6 0 0 1 12 0 v14 h-12 z" class="s"/>
    <line x1="40" y1="172" x2="160" y2="172" class="s"/>
    <text x="100" y="252" class="lbl">Converted building</text>
    <text x="100" y="270" class="sub">Former church, chapel or hall</text>
    <text x="100" y="286" class="ex">the most common in Britain</text>
  </g>
</g></svg>'''

def svg_inside():
    return '''<svg viewBox="0 0 720 300" class="dgm" role="img" aria-label="What to expect walking into a Hindu temple: leave shoes at the entrance, enter the main hall, approach the shrine where the murtis are kept, ring the bell, receive the aarti flame and prasad.">
<g class="dgm-ink">
  <rect x="24" y="46" width="672" height="176" rx="10" class="f0"/><rect x="24" y="46" width="672" height="176" rx="10" class="s"/>
  <!-- entrance -->
  <rect x="44" y="120" width="8" height="70" class="f3"/>
  <text x="92" y="86" class="num">1</text><circle cx="86" cy="80" r="13" class="s"/>
  <rect x="66" y="140" width="46" height="10" rx="2" class="f2"/><rect x="66" y="140" width="46" height="10" rx="2" class="s"/>
  <rect x="66" y="154" width="46" height="10" rx="2" class="f2"/><rect x="66" y="154" width="46" height="10" rx="2" class="s"/>
  <text x="89" y="184" class="tag">Shoe rack</text><text x="89" y="198" class="tag2">shoes off here</text>
  <!-- hall -->
  <text x="270" y="86" class="num">2</text><circle cx="264" cy="80" r="13" class="s"/>
  <g class="f2"><circle cx="212" cy="150" r="9"/><circle cx="248" cy="162" r="9"/><circle cx="288" cy="148" r="9"/><circle cx="322" cy="164" r="9"/></g>
  <g class="s"><circle cx="212" cy="150" r="9"/><circle cx="248" cy="162" r="9"/><circle cx="288" cy="148" r="9"/><circle cx="322" cy="164" r="9"/></g>
  <text x="266" y="192" class="tag">Main hall</text><text x="266" y="206" class="tag2">people sit on the floor</text>
  <!-- bell -->
  <path d="M400 96 v10 M392 122 a8 14 0 0 1 16 0 z" class="f3"/><path d="M392 122 a8 14 0 0 1 16 0 z" class="s"/>
  <text x="400" y="140" class="tag2">bell</text>
  <!-- shrine -->
  <text x="556" y="86" class="num">3</text><circle cx="550" cy="80" r="13" class="s"/>
  <path d="M498 178 L498 116 q52 -34 104 0 L602 178 Z" class="f1"/><path d="M498 178 L498 116 q52 -34 104 0 L602 178 Z" class="s"/>
  <rect x="522" y="132" width="20" height="46" rx="9" class="f3"/><rect x="558" y="132" width="20" height="46" rx="9" class="f3"/>
  <path d="M620 168 q10 -18 20 0 a10 10 0 0 1 -20 0 z" class="f3"/>
  <text x="640" y="150" class="tag2">aarti lamp</text>
  <text x="550" y="196" class="tag">Shrine (garbhagriha)</text><text x="550" y="210" class="tag2">murtis are kept here</text>
  <path d="M132 168 L186 168" class="arw"/><path d="M348 168 L470 168" class="arw"/>
  <text x="360" y="264" class="cap">Leave shoes at the door · sit or stand quietly in the hall · approach the shrine, receive the flame and prasad</text>
</g></svg>'''

ETIQUETTE = [
 ("Shoes come off", "Always, without exception. There will be racks or shelves at the entrance. Socks are fine. This is about keeping the shrine space clean, not about you personally."),
 ("Dress modestly", "Cover shoulders and knees. No special clothing is needed and you don't need to wear Indian dress. Avoid clothing with leather where you can \u2014 some temples ask visitors to leave leather belts and bags outside."),
 ("You can go in", "Hindu temples are open to people of any faith or none. You do not need to be Hindu, and you will not be asked to worship or to convert. Standing quietly and observing is entirely normal and welcome."),
 ("Photography \u2014 ask first", "Rules vary a lot. Many temples allow photos of the building but not of the murtis or during worship. Some prohibit phones in the shrine entirely. Always ask before taking a picture."),
 ("Prasad will be offered", "Blessed food \u2014 often fruit, sweets or a full vegetarian meal \u2014 is shared with everyone. Accept it with your right hand if you can. It is a gesture of hospitality; declining politely is fine if you have allergies."),
 ("Don't point your feet at the shrine", "When sitting on the floor, cross your legs or tuck your feet under you rather than stretching them toward the deities."),
 ("Walk clockwise", "If people are circling the shrine (pradakshina), follow the same direction."),
 ("Donations are voluntary", "There is usually a donation box (hundi). Nobody will ask you for money, and there is no entry fee."),
 ("Food is vegetarian", "Temple kitchens are vegetarian, and many exclude onion and garlic. Do not bring meat, alcohol or leather items into the building."),
 ("Menstruation", "Some temples ask women not to enter the shrine area during menstruation. Practice varies widely and many British mandirs do not raise it at all. If it matters to you, ask the temple when you contact them."),
]

TEACHER_QS = [
 "Do you welcome school groups, and what size can you accommodate?",
 "Is there a quiet time of day that would suit a class visit?",
 "Is there coach or minibus parking nearby?",
 "Is the building step-free, and are there accessible toilets?",
 "Will someone be available to speak to the pupils, and in English?",
 "Are photographs allowed, and can pupils take notes inside?",
 "Should pupils bring anything, or avoid bringing anything?",
 "Is there a donation you would suggest for a group visit?",
]

def build_visiting():
    et = "".join('<div class="et-card"><h3>%s</h3><p>%s</p></div>' % (e(a), e(b)) for a, b in ETIQUETTE)
    qs = "".join("<li>%s</li>" % e(q) for q in TEACHER_QS)
    cities = "".join(
        '<a class="city-card" href="%s.html"><b>%s</b><span>%d temple%s</span></a>'
        % (slug, e(nm), len(BY_CITY[slug]), "" if len(BY_CITY[slug]) == 1 else "s")
        for nm, slug, la, lo, rad, z, bl in CITIES if BY_CITY[slug])
    faqs = [
     ("Can non-Hindus visit a Hindu temple?",
      "Yes. Hindu temples in Britain are open to visitors of any faith or none. You will not be asked to take part in worship or to convert. Remove your shoes, dress modestly, and be quiet and respectful \u2014 that is all that is expected."),
     ("Do I need to book to visit a Hindu temple?",
      "As an individual, usually not \u2014 just arrive during opening hours. For a group or a school class, always contact the temple first. Use the <a href=\"index.html\">temple finder</a> to get their phone number or website."),
     ("What should I wear to a Hindu temple?",
      "Ordinary modest clothing covering shoulders and knees. No Indian dress is required. Shoes always come off at the entrance, so wear socks if you would rather not be barefoot."),
     ("Are Hindu temples free to enter?",
      "Yes. There is no entry charge. Most temples have a donation box, but donating is entirely voluntary."),
     ("Can I take photographs inside?",
      "Ask first. Policies differ: many temples allow photos of the building but not of the murtis or during worship, and some do not allow phones in the shrine at all."),
     ("How do I arrange a school visit to a Hindu temple?",
      "Contact the temple directly \u2014 this directory is not a booking service. Find your nearest temples with the <a href=\"index.html\">postcode finder</a>, then phone or email them using the details on their listing. The questions to ask are listed above."),
    ]
    faq_html = "".join("<details><summary>%s</summary><p>%s</p></details>" % (e(q), a) for q, a in faqs)
    import re as _re
    faq_ld = [{"@type": "Question", "name": q,
               "acceptedAnswer": {"@type": "Answer", "text": _re.sub(r"<[^>]+>", "", a)}} for q, a in faqs]

    body = sub("""
<div class="page wrap">
  <nav class="crumb" aria-label="Breadcrumb"><a href="index.html">Home</a> › Visiting a temple</nav>
  <h1>Visiting a Hindu temple in the UK</h1>
  <div class="prose">
    <p class="lede">A practical guide for first-time visitors, teachers planning an RE trip, and anyone who has walked past a mandir and wondered whether they could go in. <b>You can.</b> Hindu temples in Britain are open to people of any faith or none, and there is no entry charge.</p>
    <p class="note-box">This site is a directory, not a booking service \u2014 it is run by one person, and visits are arranged with each temple directly. What follows is everything worth knowing before you make that call.</p>

    <h2>What the building will look like</h2>
    <p>Britain's mandirs fall into three broad architectural types. Most are the third \u2014 ordinary buildings, often former churches and chapels, converted by communities who arrived with very little. A converted hall is no less a temple: a building becomes a mandir when the murtis are installed in it.</p>
  </div>
  __SVG1__
  <div class="prose">
    <p class="dgm-note">Which type you find near you depends on the community that built it. The <a href="temple-traditions.html">five traditions guide</a> explains the differences, and every listing in this directory is labelled by tradition.</p>

    <h2>What happens when you walk in</h2>
  </div>
  __SVG2__
  <div class="prose">
    <p class="dgm-note">You will not be expected to do anything beyond removing your shoes and being quiet. Standing at the back and watching is completely normal.</p>
    <h2>The things worth knowing</h2>
  </div>
  <div class="et-grid">__ET__</div>

  <div class="prose">
    <h2 id="schools">For teachers and group leaders</h2>
    <p>A mandir visit is one of the easiest ways to make RE concrete, and most temples are glad to host schools \u2014 but they are run by volunteers, so give plenty of notice and always ring ahead. <b>Contact the temple directly; this site cannot arrange it for you.</b></p>
    <p>Use the <a href="index.html">postcode finder</a> to see which temples are nearest your school, then use the phone number or website on their listing. Worth asking when you call:</p>
    <ul class="ask-list">__QS__</ul>
    <p>Two things that make a visit go well: tell pupils in advance that shoes come off and phones may need to stay in bags, and ask the temple what their own community would most like the class to understand. The answer is often not what a textbook would predict.</p>
    <p>It also helps to know what kind of temple you are visiting. A Tamil kovil, a BAPS mandir and a community Sanatan temple offer very different experiences \u2014 see the <a href="temple-traditions.html">guide to the five traditions</a>.</p>

    <h2>Find temples near you</h2>
    <p>Search any UK postcode on the <a href="index.html">interactive map</a>, or start from a city:</p>
  </div>
  <div class="city-grid">__CITIES__</div>
  <div class="prose" style="margin-top:20px">
    <p><a href="directory.html">Browse all __N__ temples by region \u2192</a></p>
    <h2>Common questions</h2>
  </div>
  <div class="faq">__FAQ__</div>
</div>""", SVG1=svg_inside_wrap(svg_temple_types(), "The three main temple building types found in Britain."),
           SVG2=svg_inside_wrap(svg_inside(), "A typical first visit, from the door to the shrine."),
           ET=et, QS=qs, CITIES=cities, N=str(N), FAQ=faq_html)

    page = (head("Visiting a Hindu Temple in the UK — A Guide for Schools & Visitors",
        "What to expect on your first visit to a Hindu mandir: shoes, dress, photography, prasad and etiquette \u2014 plus how teachers can arrange a school visit and find temples nearby.",
        BASE + "/visiting-a-hindu-temple.html", active="visiting",
        extra="\n".join([
            ld({"@context": "https://schema.org", "@type": "Article",
                "headline": "Visiting a Hindu temple in the UK",
                "description": "A practical guide to visiting a Hindu mandir in Britain, for first-time visitors and school groups.",
                "author": {"@type": "Organization", "name": "Hindu Temples UK"},
                "publisher": {"@type": "Organization", "name": "Hindu Temples UK",
                              "logo": {"@type": "ImageObject", "url": BASE + "/favicon.svg"}},
                "datePublished": TODAY, "dateModified": TODAY,
                "mainEntityOfPage": BASE + "/visiting-a-hindu-temple.html"}),
            ld({"@context": "https://schema.org", "@type": "FAQPage", "mainEntity": faq_ld}),
            ld(ld_breadcrumb([("Home", ""), ("Visiting a temple", "visiting-a-hindu-temple.html")]))]))
        + header_html("visiting") + body + footer_html())
    open("visiting-a-hindu-temple.html", "w", encoding="utf-8").write(page)


def svg_inside_wrap(svg, caption):
    return '<figure class="dgm-fig">%s<figcaption>%s</figcaption></figure>' % (svg, e(caption))


def build_traditions():
    cards = []
    for code, title, what, see, examples in TRADITIONS:
        n = sum(1 for t in T if t["trad"] == code)
        ex = "".join("<li>%s</li>" % e(x) for x in examples)
        cards.append(
          '<section class="trad-card" id="%s">'
          '<div class="trad-head"><span class="trad-swatch" style="background:%s"></span>'
          '<h2>%s</h2><span class="trad-count">%d in this directory</span></div>'
          '<p>%s</p>'
          '<h3>What you\u2019ll notice</h3><p>%s</p>'
          '<h3>Examples</h3><ul class="trad-ex">%s</ul>'
          '<p class="trad-link"><a href="index.html?trad=%s">Show all %d on the map \u2192</a></p>'
          '</section>' % (code.lower(), TRAD_COLOR[code], e(title), n, e(what), e(see), ex, code, n))

    body = sub("""
<div class="page wrap">
  <nav class="crumb" aria-label="Breadcrumb"><a href="index.html">Home</a> › Temple traditions</nav>
  <h1>The five kinds of Hindu temple in Britain</h1>
  <div class="prose">
    <p class="lede">Not all mandirs are alike. A Tamil kovil in Croydon and a BAPS mandir in Neasden are both Hindu temples, but they look different, sound different and worship differently. Every temple in this directory is labelled by tradition \u2014 here is what those labels mean.</p>
    <p>These are not sects in competition, and most Hindus move between them freely. They reflect where in the Hindu world a community came from, and which lineage it follows. If you are visiting for the first time, knowing which type you are entering tells you a great deal about what you will see.</p>
  </div>
  __CARDS__
  <div class="prose">
    <p class="backlink"><a href="visiting-a-hindu-temple.html">What to expect when you visit \u2192</a><br>
    <a href="directory.html">Browse all __N__ temples by region \u2192</a></p>
  </div>
</div>""", CARDS="".join(cards), N=str(N))

    page = (head("The Five Kinds of Hindu Temple in Britain — Traditions Explained",
        "Saiva kovils, BAPS mandirs, ISKCON temples, Swaminarayan and Sanatan community mandirs \u2014 what the differences are, what you\u2019ll see in each, and which UK temples belong to which.",
        BASE + "/temple-traditions.html", active="",
        extra="\n".join([
            ld({"@context": "https://schema.org", "@type": "Article",
                "headline": "The five kinds of Hindu temple in Britain",
                "author": {"@type": "Organization", "name": "Hindu Temples UK"},
                "publisher": {"@type": "Organization", "name": "Hindu Temples UK",
                              "logo": {"@type": "ImageObject", "url": BASE + "/favicon.svg"}},
                "datePublished": TODAY, "dateModified": TODAY,
                "mainEntityOfPage": BASE + "/temple-traditions.html"}),
            ld(ld_breadcrumb([("Home", ""), ("Temple traditions", "temple-traditions.html")]))]))
        + header_html("") + body + footer_html())
    open("temple-traditions.html", "w", encoding="utf-8").write(page)

# ---------------------------------------------------------------- community pages
COMMUNITY = [
 {"name": "Cambridge", "slug": "cambridge", "lat": 52.2053, "lng": 0.1218,
  "intro": "Cambridge has a large and fast-growing Hindu community \u2014 and no Hindu temple. "
           "The nearest mandir is over 25 miles away, so families here celebrate Diwali, Navratri, "
           "Ganesh Chaturthi and Holi in hired halls and community centres, organised by volunteers. "
           "A campaign led by <b>Hindu Samaj Northstowe</b>, a registered charity, is working to change that "
           "and secure a permanent mandir for Cambridgeshire. Cambridge needs a temple; until it has one, "
           "this page gathers the festival events happening across the city in one place, so nobody has to "
           "miss out simply because the information was scattered.",
  "campaign": ("A Hindu Temple for Cambridge", "https://cambridgetemple.org",
               "Hindu Samaj Northstowe \u00b7 registered charity 1213652"),
  "links": [("Hindu Samaj Northstowe", "https://www.hindusamajnorthstowe.org/"),
            ("Cambridge Hindu Temple campaign", "https://cambridgetemple.org"),
            ("Campaign on Facebook", "https://www.facebook.com/CambridgeHinduTemple")]},
]

def load_events():
    if not os.path.exists("temples.xlsx"): return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    wb = load_workbook("temples.xlsx", data_only=True)
    if "Events" not in wb.sheetnames: return []
    ws = wb["Events"]; rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    hdr = [s(h) for h in rows[0]]
    ix = {k: (hdr.index(k) if k in hdr else None) for k in
          ["Date","End date","Time","Title","City","Venue","Address","Organiser",
           "Link","Poster","Details","Free?","Source","Checked"]}
    def cell(row, k):
        i = ix[k]; v = row[i] if i is not None and i < len(row) else None
        if hasattr(v, "strftime"): return v.strftime("%Y-%m-%d")
        return s(v)
    out = []
    for row in rows[1:]:
        d = cell(row, "Date"); t = cell(row, "Title")
        if not d or not t: continue
        end = cell(row, "End date") or d
        if end < TODAY: continue
        out.append({k.lower().replace(" ", "").replace("?", ""): cell(row, k) for k in ix})
        out[-1].update({"date": d, "end": end, "title": t})
    out.sort(key=lambda x: (x["date"], x["title"]))
    return out

def month_grid(year, month, marks):
    """marks: {'YYYY-MM-DD': ('local'|'fest', label)}"""
    import calendar as _c, datetime as _dt
    _c.setfirstweekday(_c.MONDAY)
    name = _dt.date(year, month, 1).strftime("%B %Y")
    head = "".join('<th scope="col"><abbr title="%s">%s</abbr></th>' % (d, d[0])
                   for d in ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"])
    body = []
    for week in _c.monthcalendar(year, month):
        cells = []
        for day in week:
            if day == 0:
                cells.append('<td class="pad"></td>'); continue
            iso = "%04d-%02d-%02d" % (year, month, day)
            m = marks.get(iso)
            if m:
                kind, label = m
                cells.append('<td class="on %s" data-until="%s"><a href="#e-%s" title="%s">%d</a></td>'
                             % (kind, iso, iso, e(label), day))
            elif iso == TODAY:
                cells.append('<td class="today">%d</td>' % day)
            else:
                cells.append("<td>%d</td>" % day)
        body.append("<tr>%s</tr>" % "".join(cells))
    return ('<div class="cal" data-festblock><p class="cal-m">%s</p>'
            '<table><thead><tr>%s</tr></thead><tbody>%s</tbody></table></div>'
            % (name, head, "".join(body)))

def build_community_pages():
    events = load_events()
    made = []
    for c in COMMUNITY:
        evs = [x for x in events if x["city"].strip().lower() == c["name"].lower()]
        marks = {}
        for f in FEST_NATIONAL:
            marks[f["date"]] = ("fest", f["fest"])
        for x in evs:
            marks[x["date"]] = ("local", x["title"])

        # months to show: this month through the last dated item (max 6)
        import datetime as _dt
        start = _dt.date.fromisoformat(TODAY).replace(day=1)
        last_iso = max([x["end"] for x in evs] + [f["end"] for f in FEST_NATIONAL] + [TODAY])
        last = _dt.date.fromisoformat(last_iso).replace(day=1)
        months, cur, guard = [], start, 0
        while cur <= last and guard < 6:
            months.append(month_grid(cur.year, cur.month, marks))
            cur = (cur.replace(day=28) + _dt.timedelta(days=8)).replace(day=1); guard += 1

        # local event list
        if evs:
            items = []
            for x in evs:
                bits = []
                if x["time"]: bits.append(e(x["time"]))
                if x["venue"]: bits.append(e(x["venue"]))
                links = []
                if x["link"]: links.append('<a href="%s" target="_blank" rel="noopener">Event details \u2192</a>' % e(x["link"]))
                poster = ('<img class="ev-poster" src="%s" alt="Poster for %s" loading="lazy"/>'
                          % (e(x["poster"]), e(x["title"]))) if x["poster"] else ""
                items.append(
                  '<article class="ev" id="e-%s" data-until="%s">%s'
                  '<div class="ev-body"><p class="ev-date">%s</p><h3>%s</h3>'
                  '%s%s%s%s</div></article>' % (
                    e(x["date"]), e(x["end"]), poster,
                    e(pretty_date(x["date"], x["end"])), e(x["title"]),
                    ('<p class="ev-meta">%s</p>' % " \u00b7 ".join(bits)) if bits else "",
                    ('<p class="ev-addr">%s</p>' % e(x["address"])) if x["address"] else "",
                    ('<p class="ev-det">%s</p>' % e(x["details"])) if x["details"] else "",
                    ('<p class="ev-org">Organised by <b>%s</b>%s</p>'
                     % (e(x["organiser"]), (" \u00b7 " + links[0]) if links else "")) if x["organiser"] else
                    (('<p class="ev-org">%s</p>' % links[0]) if links else "")))
            ev_html = '<div class="ev-list" data-festblock>%s</div>' % "".join(items)
        else:
            ev_html = ('<div class="empty" style="margin-top:18px"><p class="big">Events are being added</p>'
                       '<p>Groups across %s are organising festival programmes now. '
                       'If you\u2019re organising one, <a href="%s">send us the details</a> '
                       'and it will appear here.</p></div>' % (e(c["name"]), MAILTO))

        near = sorted(T, key=lambda t: miles(c["lat"], c["lng"], t["lat"], t["lng"]))[:5]
        near_html = "".join(
          '<li><b>%s</b> <span class="dloc">%s \u00b7 %.0f miles</span>%s</li>'
          % (e(t["name"]), e(t["area"]), miles(c["lat"], c["lng"], t["lat"], t["lng"]),
             (' <a href="%s" target="_blank" rel="noopener">Website \u2197</a>' % e(t["url"])) if t.get("url") else "")
          for t in near)

        ctitle, curl, cnote = c["campaign"]
        clinks = "".join('<li><a href="%s" target="_blank" rel="noopener">%s</a></li>' % (e(u), e(n))
                         for n, u in c["links"])
        fest_html = ('<ul class="fest-list">%s</ul>' % fest_rows_html(FEST_NATIONAL)) if FEST_NATIONAL else ""

        body = sub("""
<div class="page wrap">
  <nav class="crumb" aria-label="Breadcrumb"><a href="index.html">Home</a> › __NM__</nav>
  <h1>Hindu festivals &amp; events in __NM__</h1>
  <div class="prose">
    <p class="lede">__INTRO__</p>
  </div>

  <div class="campaign-box">
    <p class="cb-h">The campaign for a Cambridge mandir</p>
    <p>__CNOTE__</p>
    <p><a class="btn btn-primary btn-sm" href="__CURL__" target="_blank" rel="noopener">__CTITLE__ →</a></p>
  </div>

  <div class="prose"><h2>What&rsquo;s on</h2>
    <p class="dgm-note">Events organised by Hindu groups across __NM__. Dates disappear from this page once they have passed. Always check with the organiser before travelling &mdash; venues and times can change.</p>
  </div>
  <div class="cal-wrap">__MONTHS__</div>
  <p class="cal-key"><span class="k local"></span> Local event &nbsp; <span class="k fest"></span> Festival date</p>
  __EVENTS__

  <div class="prose"><h2>Festival dates this season</h2>
    <p class="dgm-note">National dates. Groups in __NM__ may observe on a nearby day &mdash; check each event above.</p>
  </div>
  __FEST__

  <div class="prose"><h2>Nearest temples to __NM__</h2>
    <p>Until __NM__ has a mandir of its own, these are the closest:</p>
  </div>
  <ul class="near-list">__NEAR__</ul>
  <div class="prose">
    <p><a href="index.html?postcode=CB1">Find your nearest temple by postcode \u2192</a></p>
    <h2>Get involved</h2>
    <ul class="f-links">__CLINKS__</ul>
    <p class="dgm-note">Organising a festival event in __NM__? <a href="__MAILTO__">Send us the details</a> and we\u2019ll add it \u2014 free, and we\u2019ll credit and link your group.</p>
  </div>
</div>""", NM=e(c["name"]), INTRO=c["intro"], CNOTE=e(cnote), CURL=e(curl), CTITLE=e(ctitle),
           MONTHS="".join(months), EVENTS=ev_html, FEST=fest_html, NEAR=near_html,
           CLINKS=clinks, MAILTO=MAILTO)

        page = (head("Hindu Festivals & Events in %s — Diwali, Navratri & More" % c["name"],
            "What's on for Hindu families in %s: Diwali, Navratri, Ganesh Chaturthi and Holi events, who's organising them, and the campaign for a Cambridge mandir." % c["name"],
            "%s/%s.html" % (BASE, c["slug"]), active="",
            extra="\n".join([
                ld({"@context": "https://schema.org", "@type": "CollectionPage",
                    "name": "Hindu festivals and events in %s" % c["name"],
                    "url": "%s/%s.html" % (BASE, c["slug"])}),
                ld(ld_breadcrumb([("Home", ""), (c["name"], c["slug"] + ".html")]))]))
            + header_html("") + body + footer_html())
        open("%s.html" % c["slug"], "w", encoding="utf-8").write(page)
        made.append((c["name"], c["slug"], len(evs)))
    return made

# ---------------------------------------------------------------- assets
def build_assets(city_slugs):
    # community pages are added to the sitemap alongside city pages
    open("favicon.svg", "w", encoding="utf-8").write(FAVICON_SVG)
    json.dump({"name": "Hindu Temples UK", "short_name": "Temples UK",
               "start_url": "/", "display": "browser",
               "background_color": "#FAF6EE", "theme_color": "#BE3524",
               "icons": [{"src": "favicon.svg", "sizes": "any", "type": "image/svg+xml"}]},
              open("site.webmanifest", "w", encoding="utf-8"), indent=1)

    urls = (["", "directory.html", "temple-map.html", "visiting-a-hindu-temple.html",
             "temple-traditions.html", "hinduism-in-britain.html", "about.html"]
            + ["%s.html" % c["slug"] for c in COMMUNITY]) + ["%s.html" % s for s in city_slugs]
    pr = {"": "1.0", "directory.html": "0.9", "temple-map.html": "0.8",
          "hinduism-in-britain.html": "0.8",
          "visiting-a-hindu-temple.html": "0.9", "temple-traditions.html": "0.8", "about.html": "0.5"}
    sm = ['<?xml version="1.0" encoding="UTF-8"?>',
          '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">']
    for u in urls:
        sm.append("<url><loc>%s/%s</loc><lastmod>%s</lastmod><priority>%s</priority></url>" %
                  (BASE, u, TODAY, pr.get(u, "0.7")))
    sm.append("</urlset>")
    open("sitemap.xml", "w", encoding="utf-8").write("\n".join(sm))
    open("robots.txt", "w", encoding="utf-8").write(
        "User-agent: *\nAllow: /\nSitemap: %s/sitemap.xml\n" % BASE)

def build_images():
    """og-image.png (1200x630) + apple-touch-icon.png (180x180). Optional: skipped if PIL missing."""
    global HAS_OG
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        print("assets: PIL not installed - skipping og-image + apple-touch-icon"); return
    try:
        f_big = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf", 78)
        f_sub = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 34)
        f_tag = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf", 24)
    except OSError:
        print("assets: fonts unavailable - skipping og-image"); return

    W, H = 1200, 630
    img = Image.new("RGB", (W, H), "#FAF6EE")
    d = ImageDraw.Draw(img)
    d.rectangle([0, 0, W, 8], fill="#E8A13A")
    d.rectangle([0, 8, W, 14], fill="#BE3524")
    # temple mark
    mx, my = 92, 130
    d.rounded_rectangle([mx, my, mx + 150, my + 150], 30, fill="#BE3524")
    d.polygon([(mx + 75, my + 18), (mx + 118, my + 92), (mx + 32, my + 92)], fill="#FBF3E4")
    d.rectangle([mx + 28, my + 92, mx + 122, my + 122], fill="#FBF3E4")
    d.rectangle([mx + 64, my + 99, mx + 86, my + 122], fill="#BE3524")
    d.text((92, 330), "Hindu Temples UK", font=f_big, fill="#231B14")
    d.text((94, 432), "Find your nearest mandir — %d temples mapped" % N, font=f_sub, fill="#5F5344")
    d.text((94, 540), "hindu-temples.uk", font=f_tag, fill="#BE3524")
    img.save("og-image.png", optimize=True)

    icon = Image.new("RGB", (180, 180), "#BE3524")
    di = ImageDraw.Draw(icon)
    di.polygon([(90, 20), (128, 106), (52, 106)], fill="#FBF3E4")
    di.rectangle([44, 106, 136, 142], fill="#FBF3E4")
    di.rectangle([80, 114, 100, 142], fill="#BE3524")
    icon.save("apple-touch-icon.png", optimize=True)
    HAS_OG = True
    print("assets: og-image.png + apple-touch-icon.png generated")

# ---------------------------------------------------------------- run
def main():
    build_images()               # sets HAS_OG before pages reference it
    build_index()
    made = build_city_pages()
    build_directory()
    build_temple_map()
    comm = build_community_pages()
    build_visiting()
    build_traditions()
    build_history()
    build_about()
    build_404()
    build_legacy_redirect()
    build_assets([slug for _, slug, _ in made])
    for nm, slug, n in comm:
        print("  community: %-12s %d local event(s)" % (slug, n))
    print("pages: index + directory + temple-map + about + 404 + %d city pages" % len(made))
    for nm, slug, n in made: print("  %-14s %d" % (slug, n))
    print("done: %d temples, sitemap has %d URLs" % (N, 7 + len(made)))

if __name__ == "__main__":
    main()
