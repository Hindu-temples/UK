#!/usr/bin/env python3
"""
Rebuild temples.json (used by the website) from temples.xlsx.

Usage:
    python build_from_excel.py                # reads temples.xlsx, writes temples.json
    python build_from_excel.py my.xlsx        # custom input
    python build_from_excel.py my.xlsx out.json

Requires: pip install openpyxl
Edit the "Temples" sheet in Excel, save, then run this. Reads columns BY HEADER NAME,
so you can reorder or add extra columns without breaking anything.
"""
import sys, json, re
from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "temples.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "temples.json"

# Excel header  ->  JSON key
FIELDS = {
    "Temple Name": "name", "Area": "area", "Region": "region", "County": "county",
    "Latitude": "lat", "Longitude": "lng", "Address": "addr", "Phone": "phone",
    "Tradition": "trad", "Notes": "note", "Featured": "hl", "Status": "status",
    "Opening hours": "hours", "Website": "url", "Facebook": "fb", "Instagram": "ig",
}
TRAD_CODES = {"BAPS", "ISKCON", "Saiva", "Swaminarayan", "General"}
TRAD_LABELS = {  # accept friendly labels too, map back to code
    "baps swaminarayan": "BAPS", "iskcon / vaishnava": "ISKCON",
    "saiva / tamil": "Saiva", "swaminarayan (other)": "Swaminarayan",
    "sanatan / community": "General",
}
CORE = ["name", "area", "region", "county", "lat", "lng", "addr", "phone", "trad", "note", "hl", "approx"]
OPTIONAL = ["status", "hours", "url", "fb", "ig"]

def s(v):  # clean string cell
    return "" if v is None else str(v).strip()

def num(v):
    if v in (None, ""): return None
    try: return round(float(v), 4)
    except (TypeError, ValueError): return None

wb = load_workbook(SRC, data_only=True)
ws = wb["Temples"] if "Temples" in wb.sheetnames else wb.worksheets[0]
rows = list(ws.iter_rows(values_only=True))
header = [s(h) for h in rows[0]]
idx = {name: header.index(name) for name in FIELDS if name in header}

missing_headers = [h for h in ("Temple Name", "Latitude", "Longitude") if h not in idx]
if missing_headers:
    sys.exit(f"ERROR: required column(s) not found: {missing_headers}")

def get(row, header_name):
    i = idx.get(header_name)
    return row[i] if i is not None and i < len(row) else None

out, skipped = [], []
for r, row in enumerate(rows[1:], start=2):
    name = s(get(row, "Temple Name"))
    if not name:
        continue  # blank row
    lat, lng = num(get(row, "Latitude")), num(get(row, "Longitude"))
    if lat is None or lng is None:
        skipped.append((r, name, "missing/invalid Latitude or Longitude"))
        continue

    trad_raw = s(get(row, "Tradition"))
    trad = trad_raw if trad_raw in TRAD_CODES else TRAD_LABELS.get(trad_raw.lower(), "General")

    rec = {
        "name": name, "area": s(get(row, "Area")), "region": s(get(row, "Region")),
        "county": s(get(row, "County")), "lat": lat, "lng": lng,
        "addr": s(get(row, "Address")), "phone": s(get(row, "Phone")),
        "trad": trad, "note": s(get(row, "Notes")),
        "hl": s(get(row, "Featured")).lower() in ("yes", "y", "true", "1"),
        "approx": False,
    }
    # carry optional fields only when filled in (keeps the file lean)
    for header_name, key in (("Status", "status"), ("Opening hours", "hours"),
                             ("Website", "url"), ("Facebook", "fb"), ("Instagram", "ig")):
        val = s(get(row, header_name))
        if val and not (key == "status" and val.lower() == "open"):
            rec[key] = val
    out.append(rec)

with open(OUT, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, separators=(",", ":"))

print(f"OK  {len(out)} temples  ->  {OUT}")
if skipped:
    print(f"\nSkipped {len(skipped)} row(s) with no coordinates:")
    for r, name, why in skipped:
        print(f"  row {r}: {name}  ({why})")
